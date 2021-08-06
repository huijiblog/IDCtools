# -*- coding:utf-8 -*-

import paramiko
import time
import datetime
import os
import sqlite3
import re
from configobj import ConfigObj
import pypinyin
import openpyxl
from openpyxl.styles import Font, Alignment

#菜单
def handleMenu():
    userInput = input("\n请输入序列号或指令('quit'或'q'可返回菜单)：")
    return userInput

#管理和查看本地数据库
def manageLocalDB():
    ColorLogDecorator().active()
    if os.path.exists("alldata.db") == False:
        print(ColorLogDecorator().yellow("【WARNING】未发现本地数据库，正在初始化中...","strong"))
        time.sleep(1)
        conn = sqlite3.connect('alldata.db')
        c = conn.cursor()
        c.execute('''CREATE TABLE DEVICES
        (ID INTEGER PRIMARY KEY autoincrement,
        NAME varchar(255) NOT NULL,
        IP varchar(255) NOT NULL,
        TYPE varchar(255) NOT NULL,
        IDC varchar(255) NOT NULL);
        ''')
        c.execute("INSERT INTO DEVICES (NAME,IP,TYPE,IDC) \
        VALUES ('IDC-example-NE40-RT01','1.1.1.1','华为','测试')")
        conn.commit()
        conn.close()
        print(ColorLogDecorator().green("【INFO】数据库初始化完成","strong"))
    else:
        conn = sqlite3.connect('alldata.db')
        c = conn.cursor()
        print(ColorLogDecorator().green("【INFO】数据库连接成功！","strong"))
        selResult = c.execute("SELECT * FROM devices")
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        print("       【ID】  ----  【设备名称】  ----  【ip】  ----  【厂商】  ----  【IDC名称】 \n")
        for row in selResult:
            print('        ',row[0],'----',row[1],'----',row[2],'----',row[3],'----',row[4])

        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        conn.commit()
        conn.close()
        print("\n\n")
        while True:

            manaInput = input("【输入序列号 1.添加设备信息(单个录入) 2.批量添加设备信息 3.修改设备信息 4.删除设备 5.管理封堵设备】\n返回上一级请输入指令'quit'或'q'\n:")
            if manaInput == "1":
                # ColorLogDecorator().active()
                devInfo = []
                os.system('cls')
                print("======================================")
                print("             单个设备录入")
                print("当前进度：")
                print(ColorLogDecorator().blue("> *设备名称", "strong"))
                print("*设备IP\n*厂商名称\n*IDC名称")
                print("======================================")
                addInput = input("(取消保存并返回上一级菜单输入'q'或'quit')\n请输入信息：")
                if addInput == "":
                    print(ColorLogDecorator().red("【WARNING】不允许输入空值！正在跳转到上一级目录...", "strong"))
                    os.system('pause')
                    os.system('cls')
                    continue
                elif addInput == "q" or addInput == "quit":
                    os.system('cls')
                    continue
                else:
                    devInfo.append(addInput)

                os.system('cls')
                print("======================================")
                print("              单个设备录入")
                print("当前进度：")
                print("*设备名称")
                print(ColorLogDecorator().blue("> *设备IP", "strong"))
                print("*厂商名称\n*IDC名称")
                print("======================================")
                addInput = input("(取消保存并返回上一级菜单输入'q'或'quit')\n(重新输入上一步请输入'c')\n请输入信息：")
                if addInput == "":
                    print(ColorLogDecorator().red("【WARNING】不允许输入空值！正在跳转到上一级目录..."))
                    time.sleep(1.5)
                    continue
                elif addInput == "q" or addInput == "quit":
                    os.system('cls')
                    continue
                elif addInput == 'c':
                    os.system('cls')
                    print("======================================")
                    print("              单个设备录入")
                    print("当前进度：")
                    print(ColorLogDecorator().blue("> *设备名称", "strong"))
                    print("*设备IP\n*厂商名称\n*IDC名称")
                    print("======================================")
                    addInput = input("(取消保存并返回上一级菜单输入'q'或'quit')\n请输入信息：")
                    if addInput == "":
                        print(ColorLogDecorator().red("【WARNING】不允许输入空值！正在跳转到上一级目录..."))
                        time.sleep(1.5)
                        os.system('cls')
                        continue
                    elif addInput == "q" or addInput == "quit":
                        os.system('cls')
                        continue
                    else:
                        devInfo[0] = addInput
                        os.system('cls')
                        print("======================================")
                        print("              单个设备录入")
                        print("当前进度：")
                        print("*设备名称")
                        print(ColorLogDecorator().blue("> *设备IP", "strong"))
                        print("*厂商名称\n*IDC名称")
                        print("======================================")
                        addInput = input("(取消保存并返回上一级菜单输入'q'或'quit')\n请输入信息：")
                        if addInput == "":
                            print(ColorLogDecorator().yellow("【WARNING】不允许输入空值！正在跳转到上一级目录..."))
                            time.sleep(1.5)
                            os.system('cls')
                            continue
                        elif addInput == "q" or addInput == "quit":
                            os.system('cls')
                            continue
                        else:
                            devInfo.append(addInput)

                else:
                    devInfo.append(addInput)

                os.system('cls')
                print("======================================")
                print("              单个设备录入")
                print("当前进度：")
                print("*设备名称")
                print("*设备IP")
                print(ColorLogDecorator().blue("> *厂商名称", "strong"))
                print("*IDC名称")
                print("======================================")
                addInput = input("(取消保存并返回上一级菜单输入'q'或'quit')\n(重新输入上一步请输入'c')\n请输入信息：")
                if addInput == "":
                    print(ColorLogDecorator().red("【WARNING】不允许输入空值！", "strong"))
                    os.system('pause')
                    os.system('cls')
                    continue
                elif addInput == "q" or addInput == "quit":
                    os.system('cls')
                    continue
                elif addInput == 'c':
                    os.system('cls')
                    print("======================================")
                    print("              单个设备录入")
                    print("当前进度：")
                    print("*设备名称")
                    print(ColorLogDecorator().blue("> *设备IP", "strong"))
                    print("*厂商名称\n*IDC名称")
                    print("======================================")
                    addInput = input("(取消保存并返回上一级菜单输入'q'或'quit')\n请输入信息：")
                    if addInput == "":
                        print(ColorLogDecorator().yellow("【WARNING】不允许输入空值！正在跳转到上一级目录..."))
                        time.sleep(1.5)
                        os.system('cls')
                        continue
                    elif addInput == "q" or addInput == "quit":
                        os.system('cls')
                        continue
                    else:
                        devInfo[1] = addInput
                        os.system('cls')
                        print("======================================")
                        print("              单个设备录入")
                        print("当前进度：")
                        print("*设备名称")
                        print("*设备IP")
                        print(ColorLogDecorator().blue("> *厂商名称", "strong"))
                        print("*IDC名称")
                        print("======================================")
                        addInput = input("(取消保存并返回上一级菜单输入'q'或'quit')\n请输入信息：")
                        if addInput == "":
                            print(ColorLogDecorator().yellow("【WARNING】不允许输入空值！正在跳转到上一级目录..."))
                            time.sleep(1.5)
                            os.system('cls')
                            continue
                        elif addInput == "q" or addInput == "quit":
                            os.system('cls')
                            continue
                        else:
                            devInfo.append(addInput)
                else:
                    devInfo.append(addInput)

                os.system('cls')
                print("======================================")
                print("              单个设备录入")
                print("当前进度：")
                print("*设备名称")
                print("*设备IP")
                print("*厂商名称")
                print(ColorLogDecorator().blue("> *IDC名称", "strong"))
                print("======================================")
                addInput = input("(取消保存并返回上一级菜单输入'q'或'quit')\n(重新输入上一步请输入'c')\n请输入信息：")
                if addInput == "":
                    print(ColorLogDecorator().red("【WARNING】不允许输入空值！正在跳转到上一级目录..."))
                    time.sleep(1.5)
                    os.system('cls')
                    continue
                elif addInput == "q" or addInput == "quit":
                    os.system('cls')
                    continue
                elif addInput == 'c':
                    os.system('cls')
                    print("======================================")
                    print("              单个设备录入")
                    print("当前进度：")
                    print("*设备名称")
                    print("*设备IP")
                    print(ColorLogDecorator().blue("> *厂商名称", "strong"))
                    print("*IDC名称")
                    print("======================================")
                    addInput = input("(取消保存并返回上一级菜单输入'q'或'quit')\n请输入信息：")
                    if addInput == "":
                        print(ColorLogDecorator().red("【WARNING】不允许输入空值！正在跳转到上一级目录..."))
                        time.sleep(1.5)
                        os.system('cls')
                        continue
                    elif addInput == "q" or addInput == "quit" or addInput == 'c':
                        os.system('cls')
                        continue
                    else:
                        devInfo[2] = addInput
                        os.system('cls')
                        print("======================================")
                        print("              单个设备录入")
                        print("当前进度：")
                        print("*设备名称")
                        print("*设备IP")
                        print("*厂商名称")
                        print(ColorLogDecorator().blue("> *IDC名称", "strong"))
                        print("======================================")
                        addInput = input("(取消保存并返回上一级菜单输入'q'或'quit')\n请输入信息：")
                        if addInput == "":
                            print(ColorLogDecorator().red("【WARNING】不允许输入空值！正在跳转到上一级目录..."))
                            time.sleep(1.5)
                            os.system('cls')
                            continue
                        elif addInput == "q" or addInput == "quit":
                            os.system('cls')
                            continue
                        else:
                            devInfo.append(addInput)
                else:
                    devInfo.append(addInput)

                print(devInfo)
                isWantSave = input(ColorLogDecorator().yellow("！！！是否要保存？ [Y/N/C]"))
                if isWantSave.upper() == "N" or isWantSave.upper() == "C":
                    print(ColorLogDecorator().red("【INFO】已取消保存)"))
                    time.sleep(1.5)
                    continue
                elif isWantSave.upper() == "Y":
                    conn = sqlite3.connect('alldata.db')
                    c = conn.cursor()
                    sql = '''INSERT INTO DEVICES (NAME,IP,TYPE,IDC) VALUES ('{0}','{1}','{2}','{3}');'''.format(devInfo[0],devInfo[1],devInfo[2],devInfo[3])
                    print(sql)
                    c.execute(sql)
                    conn.commit()
                    conn.close()
                    print(ColorLogDecorator().green("【INFO】保存成功！)"))
                    time.sleep(1.5)
                    os.system('cls')
                    continue

            elif manaInput == "2":
                if os.path.isfile('设备表.xlsx') == False:
                    wb = openpyxl.Workbook()
                    wb.create_sheet(index=0, title='Sheet1')
                    sheet = wb.worksheets[0]
                    font = Font(name='黑体', size=10, bold=True)
                    alignment = Alignment(horizontal='left',vertical='center')
                    sheet['A1'].value = '设备名称'
                    sheet['A1'].font = font
                    sheet['A1'].alignment = alignment
                    sheet['B1'].value = 'IP'
                    sheet['B1'].font = font
                    sheet['B1'].alignment = alignment
                    sheet['C1'].value = '设备厂商'
                    sheet['C1'].font = font
                    sheet['C1'].alignment = alignment
                    sheet['D1'].value = '机房'
                    sheet['D1'].font = font
                    sheet['D1'].alignment = alignment
                    wb.save('设备表.xlsx')
                    print(ColorLogDecorator().green('【INFO】检测到模板文件不存在，已自动生成！\n', 'strong'))
                isY_input= input('    请将设备信息填入模板后，在此输入Y即可批量导入：')
                if isY_input.upper() == 'Y':
                    wb = openpyxl.load_workbook('设备表.xlsx')
                    sheet = wb.worksheets[0]
                    rowsList = []
                    cellsList = []
                    for row in sheet.iter_rows():
                        for cell in row:
                            cellsList.append(cell.value)
                        rowsList.append(cellsList)
                        cellsList = []
                    if len(rowsList) < 2:
                        print(ColorLogDecorator().red('【ERROR】无法检测到数据，请先填写数据。\n', 'strong'))
                        continue
                    del rowsList[0]
                    conn = sqlite3.connect('alldata.db')
                    c = conn.cursor()
                    for devs in rowsList:
                        PreInDBvalue = c.execute("SELECT * FROM DEVICES WHERE IP = '{0}'".format(devs[1]))
                        if len(list(PreInDBvalue)) == 0:
                            c.execute("INSERT INTO DEVICES (NAME,IP,TYPE,IDC) VALUES ('{0}','{1}','{2}','{3}')".format(devs[0],devs[1],devs[2],devs[3]))
                            print(ColorLogDecorator().green("【INFO】已添加设备 {0}".format(devs[0]), "strong"))
                        else:
                            print(ColorLogDecorator().red("【ERROR】设备 {0} 已存在数据库！请勿重复添加！".format(devs[0]), "strong"))
                    conn.commit()
                    conn.close()
                    print(ColorLogDecorator().green('【INFO】批量导入完成 ！\n', 'strong'))
                else:
                    print(ColorLogDecorator().yellow('【INFO】操作已取消！\n', 'strong'))
                continue

            elif manaInput == "3":
                print('\n-----------------------------')
                IDnumInput = input("【选择要修改的设备ID号】：")
                #判断输入的值是否只由纯数字组成
                if IDnumInput.isdigit() == True:
                    sql = '''SELECT * FROM devices WHERE ID='{0}';'''.format(IDnumInput)
                    conn = sqlite3.connect('alldata.db')
                    c = conn.cursor()
                    selResult = c.execute(sql)
                    selResult1 = list(selResult)
                    conn.commit()
                    conn.close()
                    if len(selResult1) == 0:
                        print(ColorLogDecorator().red('【结果】不存在该设备\n', 'strong'))
                        continue
                    else:
                        TheDevsInfo = []
                        titleList = ['       ID：',' 设备名称：','       IP：','     厂商：','  IDC名称：']
                        #os.system('cls')
                        print('\n++++++++++++++++当前数据+++++++++++++++++\n')
                        print('{0}{1}\n'.format(titleList[0],str(selResult1[0][0])))
                        TheDevsInfo.append(str(selResult1[0][0]))
                        print('{0}{1}\n'.format(titleList[1],selResult1[0][1]))
                        TheDevsInfo.append(selResult1[0][1])
                        print('{0}{1}\n'.format(titleList[2],selResult1[0][2]))
                        TheDevsInfo.append(selResult1[0][2])
                        print('{0}{1}\n'.format(titleList[3],selResult1[0][3]))
                        TheDevsInfo.append(selResult1[0][3])
                        print('{0}{1}\n'.format(titleList[4],selResult1[0][4]))
                        TheDevsInfo.append(selResult1[0][4])
                        print('\n++++++++++++++++++++++++++++++++++++++++++++')
                        updateInput = input('修改 设备名称【不修改则留空】：')
                        if updateInput != "":
                            TheDevsInfo[1] = updateInput
                        updateInput = input('修改 IP【不修改则留空】：')
                        if updateInput != "":
                            TheDevsInfo[2] = updateInput
                        updateInput = input('修改 厂商【不修改则留空】：')
                        if updateInput != "":
                            TheDevsInfo[3] = updateInput
                        updateInput = input('修改 IDC名称【不修改则留空】：')
                        if updateInput != "":
                            TheDevsInfo[4] = updateInput

                        os.system('cls')
                        print('++++++++++++++++更新为以下数据+++++++++++++++++\n')
                        print('{0}{1}\n'.format(titleList[0], str(TheDevsInfo[0])))
                        print('{0}{1}\n'.format(titleList[1], TheDevsInfo[1]))
                        print('{0}{1}\n'.format(titleList[2], TheDevsInfo[2]))
                        print('{0}{1}\n'.format(titleList[3], TheDevsInfo[3]))
                        print('{0}{1}\n'.format(titleList[4], TheDevsInfo[4]))
                        print('\n++++++++++++++++++++++++++++++++++++++++++++')
                        updateResultInput = input(ColorLogDecorator().yellow('  是否要保存？【Y/N】 ：','bg-strong'))
                        while True:
                            if updateResultInput.upper() == 'Y':
                                sql = "UPDATE devices SET NAME='{0}',IP='{1}',TYPE='{2}',IDC='{3}' WHERE ID='{4}';".format(TheDevsInfo[1],TheDevsInfo[2],TheDevsInfo[3],TheDevsInfo[4],TheDevsInfo[0])
                                conn = sqlite3.connect('alldata.db')
                                c = conn.cursor()
                                c.execute(sql)
                                conn.commit()
                                conn.close()
                                print(ColorLogDecorator().green('\n 【INFO】保存成功！将返回主菜单...','strong'))
                                time.sleep(2)
                                break
                            elif updateResultInput.upper() == 'N':
                                break
                            else:
                                updateResultInput = input(ColorLogDecorator().yellow('  是否要保存？【Y/N】：', 'bg-strong'))
                        os.system('cls')
                        break
                else:
                    print(ColorLogDecorator().red('【WARNING】请输入正确的ID号！\n','strong'))

            elif manaInput == "4":
                print('\n------------------------------------------------------')
                delNumInput = input('   请输入要删除的设备ID号【退出请输入c或q】：')
                if delNumInput == 'c' or delNumInput == 'q' or delNumInput == 'C' or delNumInput == 'Q':
                    continue
                elif delNumInput.isdigit() == True:
                    sql = '''SELECT * FROM devices WHERE ID='{0}';'''.format(delNumInput)
                    conn = sqlite3.connect('alldata.db')
                    c = conn.cursor()
                    devsList = c.execute(sql)
                    if len(list(devsList)) != 0:
                        conn.commit()
                        conn.close()
                        while True:
                            wantDelInput = input(ColorLogDecorator().yellow('  确定要删除?【Y/N】：','strong'))
                            if wantDelInput.upper() == 'Y':
                                ###  删除设备时，如果是关联的封堵设备，也会一同从封堵设备列表中删除
                                if os.path.isfile('blockconfig.ini'):
                                    config = ConfigObj('blockconfig.ini', encoding='UTF-8')
                                    cfgDict = config['devList']
                                    for key, values in cfgDict.items():
                                        if delNumInput + ',' in values:
                                            repAfter = values.replace(delNumInput + ',', "")
                                            config['devList'][key] = repAfter
                                            config.write()
                                        elif delNumInput == values:
                                            conn = sqlite3.connect('alldata.db')
                                            c = conn.cursor()
                                            Currsql_res = c.execute(
                                                "SELECT * FROM devices WHERE ID='{0}'".format(delNumInput))
                                            Currsql_res_List = list(Currsql_res)
                                            conn.commit()
                                            conn.close()
                                            saveVal = config['list']['jifang'].replace(Currsql_res_List[0][4] + "|", "")
                                            config['list']['jifang'] = saveVal
                                            config.write()
                                            saveVal = config['list']['jifang'].replace("|" + Currsql_res_List[0][4], "")
                                            config['list']['jifang'] = saveVal
                                            config.write()
                                            saveVal = config['list']['jifang'].replace(Currsql_res_List[0][4], "")
                                            config['list']['jifang'] = saveVal
                                            config.write()
                                            del config['devList'][key]
                                            config.write()
                                        elif delNumInput in values:
                                            repAfter = values.replace(',' + delNumInput, "")
                                            config['devList'][key] = repAfter
                                            config.write()
                                ###  END

                                sql = '''DELETE FROM devices WHERE ID='{0}';'''.format(delNumInput)
                                conn = sqlite3.connect('alldata.db')
                                c = conn.cursor()
                                c.execute(sql)
                                conn.commit()
                                conn.close()
                                print(ColorLogDecorator().green('\n【INFO】删除设备成功！\n', 'strong'))
                                os.system('pause')
                                os.system('cls')
                                break
                            elif wantDelInput.upper() == 'N':
                                os.system('cls')
                                break
                            else:
                                continue
                        continue

                    else:
                        conn.commit()
                        conn.close()
                        print(ColorLogDecorator().yellow('【WARNING】不存在该设备！\n','strong'))
                        os.system('pause')
                        os.system('cls')
                        continue

                else:
                    print(ColorLogDecorator().red("【ERROR】ID号为纯数字，请正确填写！\n", "strong"))
                    os.system('pause')
                    os.system('cls')
                    continue

            elif manaInput == "5":
                os.system('cls')
                if os.path.isfile('blockconfig.ini') == False:
                    config = ConfigObj()
                    config.filename = './blockconfig.ini'
                    config['list'] = {}
                    config['list']['jifang'] = ''
                    config['devList'] = {}
                    config['account'] = {}
                    config['account']['username'] = ''
                    config['account']['password'] = ''
                    config.write()
                    print(ColorLogDecorator().green('【INFO】封堵信息库已初始化完毕，请添加封堵设备！','strong'))
                while True:
                    manaBlockInput = input("\n 【管理封堵设备】(返回上一级请输入指令'q'或'quit')\n 1.查看封堵设备列表 2.添加封堵设备 3.删除封堵设备 4.添加/更改通用设备登陆账号\n(输入)：")
                    if manaBlockInput == "1":
                        os.system('cls')
                        if os.path.isfile('blockconfig.ini'):
                            config = ConfigObj('blockconfig.ini', encoding='UTF-8')
                            if config['list']['jifang'] == '':
                                print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                                print("                                  【封堵设备列表】")
                                print(" 【ID】  ----  【设备名称】  ----  【ip】  ----  【厂商】  ----  【机房名称】 ")
                                print(
                                    "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                                os.system('pause')
                                os.system('cls')
                                continue
                            else:
                                print(
                                    "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                                print("                                  【封堵设备列表】")
                                print(" 【ID】  ----  【设备名称】  ----  【ip】  ----  【厂商】  ----  【机房名称】 ")
                                conn = sqlite3.connect('alldata.db')
                                c = conn.cursor()
                                # Currsql_res = c.execute("SELECT * FROM devices WHERE ID='{0}'".format(ipblock_Devices))
                                for key,values in config['devList'].items():
                                    if ',' in values:
                                        AfterID_Lst = values.split(',')
                                        for textAfterID in AfterID_Lst:
                                            returnsql = c.execute(
                                                "SELECT * FROM devices WHERE ID='{0}'".format(textAfterID))
                                            ressql = list(returnsql)
                                            print(
                                                "  {0}----{1}----{2}----{3}----{4}\n".format(ressql[0][0], ressql[0][1],
                                                                                             ressql[0][2], ressql[0][3],
                                                                                             ressql[0][4]))
                                    else:
                                        returnsql = c.execute("SELECT * FROM devices WHERE ID='{0}'".format(values))
                                        ressql = list(returnsql)
                                        print("  {0}----{1}----{2}----{3}----{4}\n".format(ressql[0][0], ressql[0][1], ressql[0][2], ressql[0][3], ressql[0][4]))
                                print(
                                    "+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                                conn.commit()
                                conn.close()
                                os.system('pause')
                                continue

                    elif manaBlockInput == "2":
                        os.system('cls')
                        conn = sqlite3.connect('alldata.db')
                        c = conn.cursor()
                        print(ColorLogDecorator().green("【INFO】数据库连接成功！", "strong"))
                        selResult = c.execute("SELECT * FROM devices")
                        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                        print("       【ID】  ----  【设备名称】  ----  【ip】  ----  【厂商】  ----  【IDC名称】 \n")
                        for row in selResult:
                            print(row[0], '----', row[1], '----', row[2], '----', row[3], '----', row[4])

                        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
                        conn.commit()
                        conn.close()
                        print("\n\n")
                        addValue = input('从设备信息总库导入到封堵设备列表 ，请输入要添加的设备ID(批量添加用英文逗号分隔)：')
                        if addValue == '':
                            print(ColorLogDecorator().red("【ERROR】请正确输入值！","strong"))
                            os.system('pause')
                            os.system('cls')
                            continue
                        # 批量添加
                        if ',' in addValue:
                            # 清除用户可能首尾多输入的逗号
                            addValue = addValue.strip(',')
                            # 分割判断一遍是否全为纯数字和输入是否合法，不合法则全部都跳过
                            ad_v = addValue.split(',')
                            for i in ad_v:
                                if i.isdigit() == False:
                                    print(ColorLogDecorator().red("【ERROR】请正确输入值！", "strong"))
                                    os.system('pause')
                                    continue
                            # 查询设备ID的归属机房，以写入配置文件
                            conn = sqlite3.connect('alldata.db')
                            c = conn.cursor()
                            ini_path = "./blockconfig.ini"
                            for d_id in ad_v:
                                ressql = c.execute("SELECT * FROM devices WHERE ID='{0}'".format(d_id))
                                devs_info = list(ressql)
                                if len(devs_info) > 0:
                                    config = ConfigObj(ini_path, encoding='UTF-8')
                                    beforeLst = config['list']['jifang']
                                    s = ''
                                    for single_word in pypinyin.pinyin(devs_info[0][4], style=pypinyin.NORMAL):
                                        s += ''.join(single_word)
                                    if  beforeLst != '' and devs_info[0][4] not in beforeLst:
                                        config['list']['jifang'] = beforeLst + '|' + devs_info[0][4]
                                        config['devList'][s] = d_id
                                        config.write()
                                    elif beforeLst != '' and devs_info[0][4] in beforeLst:
                                        currIDC_devs = config['devList'][s] + ',' + d_id
                                        config['devList'][s] = currIDC_devs
                                        config.write()
                                    elif beforeLst == '':
                                        config['list']['jifang'] = devs_info[0][4]
                                        config['devList'][s] = d_id
                                        config.write()
                                    print(ColorLogDecorator().green("【INFO】添加完成！","strong"))
                                else:
                                    print(ColorLogDecorator().red("【ERROR】不存在该设备！无法添加","strong"))
                            conn.commit()
                            conn.close()
                            os.system('pause')
                            os.system('cls')
                            continue

                        # 单个添加
                        elif addValue.isdigit() == True:
                            # 查询设备ID的归属机房，以写入配置文件
                            conn = sqlite3.connect('alldata.db')
                            c = conn.cursor()
                            ini_path = "./blockconfig.ini"
                            ressql = c.execute("SELECT * FROM devices WHERE ID='{0}'".format(addValue))
                            devs_info = list(ressql)
                            if len(devs_info) > 0:
                                config = ConfigObj(ini_path, encoding='UTF-8')
                                beforeLst = config['list']['jifang']
                                s = ''
                                for single_word in pypinyin.pinyin(devs_info[0][4], style=pypinyin.NORMAL):
                                    s += ''.join(single_word)
                                if beforeLst != '' and devs_info[0][4] not in beforeLst:
                                    config['list']['jifang'] = beforeLst + '|' + devs_info[0][4]
                                    config['devList'][s] = addValue
                                    config.write()
                                elif beforeLst != '' and devs_info[0][4] in beforeLst:
                                    currIDC_devs = config['devList'][s] + ',' + addValue
                                    config['devList'][s] = currIDC_devs
                                    config.write()
                                elif beforeLst == '':
                                    config['list']['jifang'] = devs_info[0][4]
                                    config['devList'][s] = addValue
                                    config.write()
                                    print(ColorLogDecorator().green("【INFO】添加完成！", "strong"))
                            else:
                                print(ColorLogDecorator().red("【ERROR】不存在该设备！无法添加","strong"))
                            conn.commit()
                            conn.close()
                            os.system('pause')
                            os.system('cls')
                            continue

                        else:
                            print(ColorLogDecorator().red("【ERROR】请正确输入值！", "strong"))
                            os.system('pause')
                            os.system('cls')
                            continue

                    elif manaBlockInput == "3":
                        deldevs = input(" 请输入要删除的设备id(多个设备用英文逗号分隔;留空或输入q则直接退出)：")
                        if "," in deldevs and deldevs != "":
                            afterCutLst = deldevs.split(",")
                            config = ConfigObj('blockconfig.ini', encoding='UTF-8')
                            cfgDict = config['devList']
                            for afterCutValue in afterCutLst:
                                for key,values in cfgDict.items():
                                    if afterCutValue + ',' in values:
                                        repAfter = values.replace(afterCutValue + ',', "")
                                        config['devList'][key] = repAfter
                                        config.write()
                                    elif afterCutValue == values:
                                        conn = sqlite3.connect('alldata.db')
                                        c = conn.cursor()
                                        Currsql_res = c.execute("SELECT * FROM devices WHERE ID='{0}'".format(afterCutValue))
                                        Currsql_res_List = list(Currsql_res)
                                        conn.commit()
                                        conn.close()
                                        saveVal = config['list']['jifang'].replace(Currsql_res_List[0][4]+"|", "")
                                        config['list']['jifang'] = saveVal
                                        config.write()
                                        config['list']['jifang'].replace("|"+Currsql_res_List[0][4], "")
                                        config['list']['jifang'] = saveVal
                                        config.write()
                                        config['list']['jifang'].replace(Currsql_res_List[0][4],"")
                                        config['list']['jifang'] = saveVal
                                        config.write()
                                        del config['devList'][key]
                                        config.write()
                                    elif afterCutValue in values:
                                        repAfter = values.replace(',' + afterCutValue, "")
                                        config['devList'][key] = repAfter
                                        config.write()
                            print('【INFO】已完成删除！\n')
                            os.system('pause')
                            os.system('cls')
                            continue

                        elif deldevs.isdigit() == True:
                            config = ConfigObj('blockconfig.ini', encoding='UTF-8')
                            cfgDict = config['devList']
                            for key, values in cfgDict.items():
                                if deldevs + ',' in values:
                                    repAfter = values.replace(deldevs + ',', "")
                                    config['devList'][key] = repAfter
                                    config.write()
                                elif deldevs == values:
                                    conn = sqlite3.connect('alldata.db')
                                    c = conn.cursor()
                                    Currsql_res = c.execute(
                                        "SELECT * FROM devices WHERE ID='{0}'".format(deldevs))
                                    Currsql_res_List = list(Currsql_res)
                                    conn.commit()
                                    conn.close()
                                    saveVal = config['list']['jifang'].replace(Currsql_res_List[0][4] + "|", "")
                                    config['list']['jifang'] = saveVal
                                    config.write()
                                    config['list']['jifang'].replace("|" + Currsql_res_List[0][4], "")
                                    config['list']['jifang'] = saveVal
                                    config.write()
                                    config['list']['jifang'].replace(Currsql_res_List[0][4], "")
                                    config['list']['jifang'] = saveVal
                                    config.write()
                                    del config['devList'][key]
                                    config.write()
                                elif deldevs in values:
                                    repAfter = values.replace(',' + deldevs, "")
                                    config['devList'][key] = repAfter
                                    config.write()
                            print(ColorLogDecorator().green('【INFO】已完成删除！\n', 'strong'))
                            os.system('pause')
                            os.system('cls')
                            continue

                        elif deldevs == "" or deldevs.upper() == "Q":
                            os.system('cls')
                            continue

                    elif manaBlockInput == "4":
                        os.system('cls')
                        config = ConfigObj('blockconfig.ini', encoding='UTF-8')
                        print("============================")
                        print("  【通用设备登录账号管理】")
                        print(" 账号：{0}".format(config['account']['username']))
                        print(" 密码：{0}".format(config['account']['password']))
                        print("============================\n")
                        modifyUser = input(" 修改账号为(留空则不修改)：")
                        modifyPass = input(" 修改密码为(留空则不修改)：")
                        if modifyUser != "":
                            config['account']['username'] = modifyUser
                        if modifyPass != "":
                            config['account']['password'] = modifyPass
                        config.write()
                        print(ColorLogDecorator().green("【INFO】修改完成！\n", "strong"))
                        os.system('pause')
                        os.system('cls')

                    elif manaBlockInput.upper() == "Q" or manaBlockInput.upper() == "QUIT":
                        os.system('cls')
                        break
                    else:
                        os.system('cls')
                        continue

            elif manaInput == "quit" or manaInput == "q":
                os.system('cls')
                break
            else:
                print(ColorLogDecorator.red("指令错误！\n", "strong"))
                os.system('pause')
                os.system('cls')
                continue


# 导出所有设备配置信息
class devConfig_export:
    def __init__(self):
        self.out_filepath = ''

    def main(self):
        conn = sqlite3.connect('alldata.db')
        c = conn.cursor()
        sqlRes = c.execute("SELECT * FROM devices")
        if os.path.isfile('blockconfig.ini'):
            config = ConfigObj('blockconfig.ini', encoding='UTF-8')
            conf_un = config['account']['username']
            conf_pw = config['account']['password']
            if conf_un == '' or conf_pw == '':
                print('【ERROR】设备通用登录账号未设置,请设置后再操作！\n')
                return
        else:
            print('【ERROR】设备通用登录账号未设置,请设置后再操作！\n')
            return

        self.out_filepath = "配置文件" + (datetime.datetime.today()).strftime('%Y-%m-%d')
        if not os.path.exists(self.out_filepath):
            os.mkdir(self.out_filepath)

        for i in sqlRes:
            if '华为' in i[3]:
                self.setHuawei(i[1],i[2],conf_un,conf_pw)
            elif 'H3C' in i[3]:
                self.setH3C(i[1],i[2],conf_un,conf_pw)
            elif '中兴' in i[3]:
                self.setZhongxing(i[1],i[2],conf_un,conf_pw)
            elif '锐捷' in i[3]:
                self.setRuijie(i[1],i[2],conf_un,conf_pw)
            elif '烽火' in i[3]:
                self.setFenghuo(i[1],i[2],conf_un,conf_pw)
            else:
                print('【ERROR】设备【{0}】未导出配置，原因：程序暂不支持【{1}】厂商\n'.format(i[1], i[3]))
        conn.commit()
        conn.close()

    def setHuawei(self, devName, conf_ip, conf_uname, conf_pass):
        try:
            # 创建ssh对象
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=conf_ip, port=22, username=conf_uname, password=conf_pass)
            # 执行命令
            command = ssh.invoke_shell()  # 激活Terminal
            time.sleep(3)
            command.send("N\n")
            command.send("screen-length 0 temporary\n")
            command.send("dis cu\n")
            data = []
            while True:
                out = command.recv(65535)
                res = out.decode(encoding='UTF-8')
                data.append(res)
                if 'return' in res:
                    break

            r_time = '.\\' + self.out_filepath + '\\' +devName + '.txt'
            res = ''.join(data)
            f1 = open(r_time, 'w')
            f1.write(res)
            f1.close()
            with open(r_time, 'r') as f2:
                final_data = f2.read().replace('\n\n', '\n')
                f2.close()
            f3 = open(r_time, 'w')
            f3.write(final_data)
            f3.close()
            ssh.close()
            print('【{0}】执行完毕！'.format(devName))
        except:
            print('【ERROR】连接 [{0}] 超时！'.format(devName))

    def setH3C(self, devName, conf_ip, conf_uname, conf_pass):
        try:
            # 创建ssh对象
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=conf_ip, port=22, username=conf_uname, password=conf_pass)
            # 执行命令
            command = ssh.invoke_shell()  # 激活Terminal
            time.sleep(3)
            command.send("screen-length disable\n")
            command.send("dis cu\n")
            data = []
            while True:
                out = command.recv(65535)
                res = out.decode(encoding='UTF-8')
                data.append(res)
                if 'return' in res:
                    break

            r_time = '.\\' + self.out_filepath + '\\' + devName + '.txt'
            res = ''.join(data)
            f1 = open(r_time, 'w')
            f1.write(res)
            f1.close()
            with open(r_time, 'r') as f2:
                final_data = f2.read().replace('\n\n\n', '\n')
                f2.close()
            f3 = open(r_time, 'w')
            f3.write(final_data)
            f3.close()
            ssh.close()
            print('【{0}】执行完毕！'.format(devName))
        except:
            print('【ERROR】连接 [{0}] 超时！'.format(devName))

    def setZhongxing(self, devName, conf_ip, conf_uname, conf_pass):
        try:
            # 创建ssh对象
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=conf_ip, port=22, username=conf_uname, password=conf_pass)
            # 执行命令
            command = ssh.invoke_shell()  # 激活Terminal
            time.sleep(3)
            command.send("enable\n")
            command.send("terminal length 0\n")
            command.send("show running-config\n")
            data = []
            while True:
                out = command.recv(65535)
                res = out.decode(encoding='UTF-8')
                data.append(res)
                if '!</static>' in res:
                    break

            r_time = '.\\' + self.out_filepath + '\\' + devName + '.txt'
            res = ''.join(data)
            f1 = open(r_time, 'w')
            f1.write(res)
            f1.close()
            with open(r_time, 'r') as f2:
                final_data = f2.read().replace('\n\n', '\n')
                f2.close()
            f3 = open(r_time, 'w')
            f3.write(final_data)
            f3.close()
            ssh.close()
            print('【{0}】执行完毕！'.format(devName))
        except:
            print('【ERROR】连接 [{0}] 超时！'.format(devName))

    def setRuijie(self, devName, conf_ip, conf_uname, conf_pass):
        try:
            # 创建ssh对象
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=conf_ip, port=22, username=conf_uname, password=conf_pass)
            # 执行命令
            command = ssh.invoke_shell()  # 激活Terminal
            time.sleep(3)
            command.send("terminal length 0\n")
            command.send("show running-config\n")
            command.send("show hosts\n")
            data = []
            while True:
                out = command.recv(65535)
                res = out.decode(encoding='UTF-8')
                data.append(res)
                if "!end" in res or "Host" in res:
                    break

            r_time = '.\\' + self.out_filepath + '\\' + devName + '.txt'
            res = ''.join(data)
            f1 = open(r_time, 'w')
            f1.write(res)
            f1.close()
            with open(r_time, 'r') as f2:
                final_data = f2.read().replace('\n\n\n', '\n')
                f2.close()
            f3 = open(r_time, 'w')
            f3.write(final_data)
            f3.close()
            ssh.close()
            print('【{0}】执行完毕！'.format(devName))
        except:
            print('【ERROR】连接 [{0}] 超时！'.format(devName))


    def setFenghuo(self, devName, conf_ip, conf_uname, conf_pass):
        try:
            # 创建ssh对象
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=conf_ip, port=22, username=conf_uname, password=conf_pass)
            # 执行命令
            command = ssh.invoke_shell()  # 激活Terminal
            time.sleep(3)
            command.send("terminal length 0\n")
            command.send("show running-config\n")
            data = []
            while True:
                out = command.recv(65535)
                res = out.decode(encoding='UTF-8')
                data.append(res)
                if "!end" in res or "Host" in res:
                    break

            r_time = '.\\' + self.out_filepath + '\\' + devName + '.txt'
            res = ''.join(data)
            f1 = open(r_time, 'w')
            f1.write(res)
            f1.close()
            with open(r_time, 'r') as f2:
                final_data = f2.read().replace('\n\n', '\n')
                f2.close()
            f3 = open(r_time, 'w')
            f3.write(final_data)
            f3.close()
            ssh.close()
            print('【{0}】执行完毕！'.format(devName))
        except:
            print('【ERROR】连接 [{0}] 超时！'.format(devName))


# ip封堵 类，只需调用main函数
class ip_block:
    #初始化一些变量属性
    def __init__(self):
        self.jifang_list = []
        self.jifang_select = ''
        self.block_ip = ''
        self.CurrBlockDev_Info = []
        ColorLogDecorator.active()

    # 读取和初始化封堵设备信息
    def Init_jifang_list(self):
        if os.path.isfile('blockconfig.ini'):
            config = ConfigObj('blockconfig.ini', encoding='UTF-8')
            if config['list']['jifang'] == '':
                return False
            else:
                if '|' in config['list']['jifang']:
                    self.jifang_list = config['list']['jifang'].split('|')
                else:
                    self.jifang_list.append(config['list']['jifang'])
                return True
        else:
            return False


    #执行封堵
    def startFengdu(self):
        s = ''
        for single_word in pypinyin.pinyin(self.jifang_select, style=pypinyin.NORMAL):
            s += ''.join(single_word)
        config = ConfigObj('blockconfig.ini', encoding='UTF-8')
        ipblock_Devices = config['devList'][s]
        if ',' in ipblock_Devices:
            Curr_devList = ipblock_Devices.split(',')
            conn = sqlite3.connect('alldata.db')
            c = conn.cursor()
            self.CurrBlockDev_Info = []
            for a in Curr_devList:
                Currsql_res = c.execute("SELECT * FROM devices WHERE ID='{0}'".format(a))
                Currsql_res_List = list(Currsql_res)
                print(Currsql_res_List)
                self.CurrBlockDev_Info.append(Currsql_res_List)
                if  'HUAWEI' in  Currsql_res_List[0][3] or 'huawei' in  Currsql_res_List[0][3] or '华为' in  Currsql_res_List[0][3]:
                    if self.setHuawei():
                        print(ColorLogDecorator().green('【INFO】封堵完成！请自行Ping测是否有效', 'strong'))
                        # self.exportToFile()
                        # print(ColorLogDecorator.green('【INFO】已导出配置记录表和封堵脚本至程序运行目录下', 'strong'))
                elif 'H3C' in Currsql_res_List[0][3] or '华三' in  Currsql_res_List[0][3]:
                    if self.setH3C():
                        print(ColorLogDecorator().green('【INFO】封堵完成！请自行Ping测是否有效', 'strong'))
                        # self.exportToFile()
                elif 'RUIJIE'in Currsql_res_List[0][3] or '锐捷' in  Currsql_res_List[0][3]:
                    if self.setRuijie():
                        pass
                        #self.exportToFile()
                elif 'ZHONGXING'in Currsql_res_List[0][3] or 'ZTE' in  Currsql_res_List[0][3] or '中兴'in Currsql_res_List[0][3]:
                    if self.setZhongxing():
                        pass
                        # self.exportToFile()
                else:
                    print(ColorLogDecorator().red('【ERROR】抱歉，本程序不支持该厂商设备！', 'strong'))

            conn.commit()
            conn.close()
            os.system('pause')
        elif ipblock_Devices.isdigit() == True:
            conn = sqlite3.connect('alldata.db')
            c = conn.cursor()
            Currsql_res = c.execute("SELECT * FROM devices WHERE ID='{0}'".format(ipblock_Devices))
            Currsql_res_List = list(Currsql_res)
            if 'HUAWEI' in Currsql_res_List[0][3] or 'huawei' in Currsql_res_List[0][3] or '华为' in Currsql_res_List[0][3]:
                if self.setHuawei():
                    print(ColorLogDecorator().green('【INFO】封堵完成！请自行Ping测是否有效', 'strong'))
                    # self.exportToFile()
                    # print(ColorLogDecorator.green('【INFO】已导出配置记录表和封堵脚本至程序运行目录下', 'strong'))
            elif 'H3C' in Currsql_res_List[0][3] or '华三' in Currsql_res_List[0][3]:
                if self.setH3C():
                    print(ColorLogDecorator().green('【INFO】封堵完成！请自行Ping测是否有效', 'strong'))
                    #self.exportToFile()
            elif 'RUIJIE' in Currsql_res_List[0][3] or '锐捷' in Currsql_res_List[0][3]:
                if self.setRuijie():
                    pass
                    #self.exportToFile()
            elif 'ZHONGXING' in Currsql_res_List[0][3] or 'ZTE' in Currsql_res_List[0][3] or '中兴' in Currsql_res_List[0][3]:
                if self.setZhongxing():
                    pass
                    #self.exportToFile()
            else:
                print(ColorLogDecorator.red('【ERROR】抱歉，本程序不支持该厂商设备！', 'strong'))
            os.system('pause')
        else:
            print(ColorLogDecorator.red('【ERROR】抱歉，发生未知错误！', 'strong'))
            os.system('pause')

    #导出封堵脚本和配置记录表
    def exportToFile(self):
        pass

    def setHuawei(self):
        config = ConfigObj('blockconfig.ini', encoding='UTF-8')
        if config['account']['username'] == '' or config['account']['password'] == '':
            print(ColorLogDecorator().red('【ERROR】设备通用SSH账号密码未设置，请设置后再使用本功能！', 'strong'))
            return False
        for i in range(len(self.CurrBlockDev_Info)):
            # 创建ssh对象
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=self.CurrBlockDev_Info[i][0][2], port=22, username=config['account']['username'], password=config['account']['password'])
            # 执行命令
            command = ssh.invoke_shell()  # 激活Terminal
            time.sleep(3)
            command.send("N\n")
            command.send("sys\n")
            command.send("ip route-static {0} 255.255.255.255 NULL0\n".format(self.block_ip))
            # VRP8版本的高端设备需加commit
            command.send("commit\n")
            command.send("return\n")
            command.send("save\n")
            command.send("Y\n")
            time.sleep(3)
            # command.recv(65535) #这里可接收结果写入日志
            ssh.close()

        return True


    def setRuijie(self):
        pass

    def setH3C(self):
        config = ConfigObj('blockconfig.ini', encoding='UTF-8')
        if config['account']['username'] == '' or config['account']['password'] == '':
            print(ColorLogDecorator().red('【ERROR】设备通用SSH账号密码未设置，请设置后再使用本功能！', 'strong'))
            return False
        for i in range(len(self.CurrBlockDev_Info)):
            # 创建ssh对象
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname=self.CurrBlockDev_Info[i][0][2], port=22, username=config['account']['username'],password=config['account']['password'])
            # 执行命令
            command = ssh.invoke_shell()  # 激活Terminal
            time.sleep(3)
            command.send("sys\n")
            command.send("ip route-static {0} 255.255.255.255 NULL0\n".format(self.block_ip))
            command.send("return\n")
            command.send("save\n")
            command.send("Y\n")
            command.send("\n")
            command.send("Y\n")
            time.sleep(1)
            # command.recv(65535) #这里可接收结果写入日志
            ssh.close()
        print(ColorLogDecorator().green('【INFO】封堵完成！请自行Ping测是否有效', 'strong'))
        return True

    def setZhongxing(self):
        pass

    def setCisco(self):
        pass


    def main(self):
        if self.Init_jifang_list():
            while True:
                os.system('cls')

                print('''
                    ===============================================
                                [选择]  >>要封堵的机房<<
                ''')
                for i in range(len(self.jifang_list)):
                    print('     {0}. {1}'.format(i+1, self.jifang_list[i]))

                print('\n    (tips: 返回上一级菜单请输入q或quit)')
                print('===============================================\n')
                sel_jifang = input('请输入：')
                if sel_jifang == 'q' or sel_jifang == 'quit' or sel_jifang == 'Q':
                    break
                elif sel_jifang.isdigit() == True and int(sel_jifang) <= len(self.jifang_list) and int(sel_jifang) != 0:
                    self.jifang_select = self.jifang_list[int(sel_jifang)-1]

                print('\n\n')
                print('*********************************************')
                inBlockIP = input('       请输入要封堵的IP：')
                check_ip = re.compile('^(1\d{2}|2[0-4]\d|25[0-5]|[1-9]\d|[1-9])\.(1\d{2}|2[0-4]\d|25[0-5]|[1-9]\d|\d)\.(1\d{2}|2[0-4]\d|25[0-5]|[1-9]\d|\d)\.(1\d{2}|2[0-4]\d|25[0-5]|[1-9]\d|\d)$')
                if check_ip.match(inBlockIP):
                    self.block_ip = inBlockIP
                    self.startFengdu()
                    os.system('cls')
                    continue
                else:
                    print(ColorLogDecorator().red('【ERROR】IP不合法！请重新操作！', 'strong'))
                    os.system('pause')
                    os.system('cls')
                    continue
        else:
            print(ColorLogDecorator().red('【ERROR】ip封堵设备信息库为空，请添加相关封堵设备！', 'strong'))
            os.system('pause')

### 自定义颜色
class ColorLogDecorator:
    """
    class:  ColorLogDecorator
    Desc:   ues for decorate the string with ANSI escape code (color function)

    class_var:
        __IS_ACTIVE: whether the decorate func active
        __DEFAULT_STYLE: the default style for a color selected
        __END_CODE: the end escape code
    """

    __IS_ACTIVE = False
    __DEFAULT_STYLE = "normal"
    __COLOR_CODE = {
        "red": {
            "normal": "\033[31m",
            "strong": "\033[1;31m",
            "bg": "\033[0;37;41m",
            "bg-strong": "\033[1;37;41m",
        },
        "green": {
            "normal": "\033[32m",
            "strong": "\033[1;32m",
            "bg": "\033[0;37;42m",
            "bg-strong": "\033[1;37;42m",
        },
        "yellow": {
            "normal": "\033[33m",
            "strong": "\033[1;33m",
            "bg": "\033[0;30;43m",
            "bg-strong": "\033[1;30;43m",
        },
        "blue": {
            "normal": "\033[34m",
            "strong": "\033[1;34m",
            "bg": "\033[0;37;44m",
            "bg-strong": "\033[1;37;44m",
        },
        "black": {
            "normal": "\033[30m",
            "strong": "\033[1;30m",
            "bg": "\033[0;37;40m",
            "bg-strong": "\033[1;37;40m",
        },
        "white": {
            "normal": "\033[37m",
            "strong": "\033[1;37;0m",
            "bg": "\033[0;30;47m",
            "bg-strong": "\033[1;30;47m",
        }
    }
    __END_CODE = "\033[0m"

    @classmethod
    def active(cls) -> None:
        """
        active the color decorate function
            it will use a special menthol for windows os
        :return: None
        """
        cls.__IS_ACTIVE = True
        if os.name == "nt":
            os.system("")

    @classmethod
    def deactivate(cls) -> None:
        """
        deactivate the color decorate function
        :return: None
        """
        cls.__IS_ACTIVE = False

    @classmethod
    def __msg_decorator(cls, msg: str, color: str, style: str) -> str:
        """
        use to decorate the msg str with special style color escape code
        :param msg: the msg str
        :param color: the color str to select
        :param style: the style str to select
        :return: decorated str
        """
        if not cls.__IS_ACTIVE:
            return msg

        style_selected = cls.__DEFAULT_STYLE if style.lower() not in cls.__COLOR_CODE[color].keys() \
            else style.lower()

        return cls.__COLOR_CODE[color][style_selected] + msg + cls.__END_CODE

    @classmethod
    def red(cls, msg: str, style: str = "normal") -> str:
        """
        red log str
        :param msg: the msg str
        :param style: the style to select
        :return: decorated str
        """
        return cls.__msg_decorator(msg, "red", style)

    @classmethod
    def green(cls, msg: str, style: str = "normal") -> str:
        """
        green log str
        :param msg: the msg str
        :param style: the style to select
        :return: decorated str
        """
        return cls.__msg_decorator(msg, "green", style)

    @classmethod
    def yellow(cls, msg: str, style: str = "normal") -> str:
        """
        yellow log str
        :param msg: the msg str
        :param style: the style to select
        :return: decorated str
        """
        return cls.__msg_decorator(msg, "yellow", style)

    @classmethod
    def blue(cls, msg: str, style: str = "normal") -> str:
        """
        blue log str
        :param msg: the msg str
        :param style: the style to select
        :return: decorated str
        """
        return cls.__msg_decorator(msg, "blue", style)

    @classmethod
    def black(cls, msg: str, style: str = "normal") -> str:
        """
        black log str
        :param msg: the msg str
        :param style: the style to select
        :return: decorated str
        """
        return cls.__msg_decorator(msg, "black", style)

    @classmethod
    def white(cls, msg: str, style: str = "normal") -> str:
        """
        white log str
        :param msg: the msg str
        :param style: the style to select
        :return: decorated str
        """
        return cls.__msg_decorator(msg, "white", style)


if __name__ == "__main__":
    menu = """
                    +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
                    +                       简易运维自动化工具-测试版                         +
                    +                                                                         +
                    +          1. 导出所有网络设备配置信息                                    +
                    +                                                                         +
                    +          2. Ping测ip地址（暂未开发）                                    +
                    +                                                                         +
                    +          3. 封堵ip地址                                                  +
                    +                                                                         +
                    +          4. 管理和查看设备信息数据库                                    +
                    +                                                                         +
                    + -----------------------------                                           +
                    + 本工具使用帮助 请输入 help                                              +
                    + 作者信息 请输入 author                                                  +
                    +                                                                         +
                    +                                 Powered By HuiJi                        +
                    +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
                """

    while True:
        print(menu)
        currInput = handleMenu()
        if currInput == "quit" or currInput == "q":
            os.system('cls')
            continue
        elif currInput == "1":
            devOut = devConfig_export()
            devOut.main()
            os.system('pause')
            os.system('cls')
            continue
        elif currInput == "2":
            os.system('cls')
            ColorLogDecorator().active()
            print(ColorLogDecorator().blue('\n【INFO】功能待开发！', 'strong'))
            os.system('pause')
            os.system('cls')
            continue
        elif currInput == "3":
            use_test = ip_block()
            use_test.main()
            #os.system('pause')
            os.system('cls')
            continue
        elif currInput == "4":
            os.system('cls')
            manageLocalDB()
            os.system('cls')
            continue
        elif currInput == "help":
            os.system('cls')
            print('''
            ----------------------------------------------------------------------------
                本工具可用于自动化运维，代替一些重复的事情，后续有时间会学习PyQT或
                其它任何可行方式，使工具改成UI界面更便于使用

                ***重点*** -----本工具完美适配Windows系统，其它操作系统下会报错

                使用说明：
                   *输入对应序列号或特定指令即可操作
                   *首次使用，请先进入[4]添加设备信息进数据库
                   *封堵功能首次使用准备步骤：添加设备信息进数据库->添加封堵设备->IP封堵
                     由于本人工作中只用到华为和H3C的设备，其它设备的封堵命令没有写进去。
                   *导出所有设备配置信息即：导出所有配置记录，目前只支持 华为、H3C、
                     锐捷、中兴、烽火 设备。
            ----------------------------------------------------------------------------
            ''')
            print("\n\n")
            os.system('pause')
            os.system('cls')
            continue
        elif currInput == "author":
            os.system('cls')
            print('''
            ------------------------------------------------
                 本人是一名网工，在工作中会有经常性的
                 重复工作，写这个工具目的除了能使工作
                 简要化之外，还能提升编程能力。
                 *作者网名：灰机
                 *具备技能：
                     Python、Linux、网络管理、路由交换、
                     PHP 等
                 *个人博客：blog.huiji888.cn

                 平时喜欢写代码、搭建和维护网站、搞搞网络，
                 希望各路大佬前来交流，互相促进。
            -------------------------------------------------
            ''')
            print("\n\n")
            os.system('pause')
            os.system('cls')
            continue

